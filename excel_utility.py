import openpyxl

def merge_excel_files(output_file, input_files):
    wb_output = openpyxl.Workbook()
    for file in input_files:
        wb_input = openpyxl.load_workbook(file)
        for sheet_name in wb_input.sheetnames:
            ws_input = wb_input[sheet_name]
            ws_output = wb_output.create_sheet(title=sheet_name)
            for row in ws_input.iter_rows():
                ws_output.append([cell.value for cell in row])
    wb_output.save(output_file)

if __name__ == "__main__":
    input_files = ["file1.xlsx", "file2.xlsx"]
    output_file = "merged.xlsx"
    merge_excel_files(output_file, input_files)